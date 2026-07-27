#!/usr/bin/env python3
"""
库位表格可视化编辑器 v6.8
作者：小刘
版本历史：
- v6.8 (2026-07-27): 右键取消锁定时同时清除选中框（cell-selecting）
- v6.7 (2026-07-27): 右键取消锁定时清除选中高亮和详情面板
- v6.6 (2026-07-27): 右键取消锁定绑定到格子上
- v6.5 (2026-07-27): 右键仅取消锁定，不触发选中
- v6.4 (2026-07-27): 右键点击取消锁定、调试日志
- v6.3 (2026-07-27): Ctrl+点击锁定详情面板、层高/承重自动保存到服务器
- v6.2 (2026-07-27): 悬停格子显示详情、清理死代码、优化批量操作提示
- v6.1 (2026-07-25): Excel 解析库从 python-calamine 改回 openpyxl，修复 CalamineWorkbook 导入错误
- v6.0 (2026-07-24): XY轴单独翻转功能上线，支持 X↔ 和 Y↕ 独立操作
- v5.x: 初始版本，基础库位编辑功能
"""
import http.server
import ssl
import json
import io
import re
import os
import socket
import time
from http import HTTPStatus

import openpyxl
from openpyxl import Workbook
import threading

PORT = 28888

# ============= SVG Icons (served from /icons/*.svg) =============

# ============= HTML Template =============
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>库位编辑器</title>
<style>
  :root {
    --bg-primary: #ffffff;
    --bg-secondary: #f8f9fa;
    --bg-tertiary: #f1f3f5;
    --bg-card: #ffffff;
    --border: rgba(0,0,0,0.08);
    --border-hover: rgba(0,0,0,0.15);
    --text-primary: #1a1a2e;
    --text-secondary: #5c5c70;
    --text-muted: #8b8b9e;
    --accent: #6366f1;
    --accent-hover: #4f46e5;
    --accent-subtle: rgba(99,102,241,0.1);
    --green: #22c55e;
    --green-subtle: rgba(34,197,94,0.1);
    --red: #ef4444;
    --red-subtle: rgba(239,68,68,0.1);
    --orange: #f59e0b;
    --orange-subtle: rgba(245,158,11,0.1);
    --shadow: 0 2px 12px rgba(0,0,0,0.08);
    --radius: 12px;
    --radius-sm: 8px;
    --transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
  }
  
  * { margin: 0; padding: 0; box-sizing: border-box; }
  
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
    background: var(--bg-primary);
    color: var(--text-primary);
    min-height: 100vh;
    font-size: 14px;
    line-height: 1.5;
    -webkit-font-smoothing: antialiased;
  }
  
  /* Header */
  .header {
    background: var(--bg-secondary);
    border-bottom: 1px solid var(--border);
    padding: 12px 24px;
    display: flex;
    align-items: center;
    gap: 12px;
    height: 56px;
  }
  
  .logo {
    display: flex;
    align-items: center;
    gap: 10px;
    color: var(--text-primary);
    text-decoration: none;
  }
  
  .logo-icon {
    width: 32px;
    height: 32px;
    background: var(--accent);
    border-radius: var(--radius-sm);
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
  }
  
  .logo-text {
    font-size: 15px;
    font-weight: 600;
    letter-spacing: -0.3px;
  }
  
  .header-actions {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-left: auto;
  }
  
  .btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 7px 14px;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    background: var(--bg-tertiary);
    color: var(--text-primary);
    font-size: 13px;
    font-weight: 500;
    cursor: pointer;
    transition: var(--transition);
    outline: none;
    text-decoration: none;
  }
  
  .btn img {
    width: 16px;
    height: 16px;
    flex-shrink: 0;
  }
  
  .btn:hover {
    background: var(--bg-card);
    border-color: var(--border-hover);
  }
  
  .btn-primary {
    background: var(--accent);
    border-color: var(--accent);
    color: white;
  }
  
  .btn-primary:hover {
    background: var(--accent-hover);
    border-color: var(--accent-hover);
  }
  
  .btn-ghost {
    background: transparent;
    border-color: transparent;
  }
  
  .btn-ghost:hover {
    background: var(--bg-tertiary);
  }
  
  .btn-icon {
    padding: 7px;
  }
  
  .btn-sm {
    padding: 5px 10px;
    font-size: 12px;
  }
  
  /* Main Layout */
  .main {
    display: flex;
    gap: 16px;
    padding: 16px;
    height: calc(100vh - 56px);
  }
  
  /* Grid Area */
  .grid-area {
    flex: 1;
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    display: flex;
    flex-direction: column;
    overflow: hidden;
    position: relative;
  }
  
  .grid-toolbar {
    display: flex;
    align-items: center;
    gap: 8px;
    padding: 12px 16px;
    border-bottom: 1px solid var(--border);
    flex-shrink: 0;
  }
  
  .z-tabs {
    display: flex;
    gap: 4px;
  }
  
  .z-tab {
    padding: 6px 14px;
    background: transparent;
    border: 1px solid transparent;
    border-radius: 6px;
    cursor: pointer;
    font-size: 13px;
    font-weight: 500;
    color: var(--text-secondary);
    transition: var(--transition);
  }
  
  .z-tab:hover {
    background: var(--bg-tertiary);
    color: var(--text-primary);
  }
  
  .z-tab.active {
    background: var(--accent-subtle);
    color: var(--accent);
    border-color: var(--accent-subtle);
  }
  
  .toolbar-spacer { flex: 1; }
  
  .zoom-controls {
    display: flex;
    align-items: center;
    gap: 4px;
    background: var(--bg-tertiary);
    border-radius: 6px;
    padding: 2px;
  }
  
  .zoom-btn {
    width: 32px;
    height: 32px;
    border: 1px solid var(--border);
    background: var(--bg-card);
    border-radius: 6px;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    transition: var(--transition);
    color: var(--text-primary);
  }
  
  .zoom-btn img {
    width: 18px;
    height: 18px;
    stroke: currentColor;
  }
  
  .zoom-btn:hover {
    background: var(--bg-card);
    color: var(--text-primary);
  }
  
  .zoom-level {
    font-size: 12px;
    color: var(--text-secondary);
    min-width: 38px;
    text-align: center;
    font-variant-numeric: tabular-nums;
  }
  
  /* Viewport */
  .grid-viewport {
    flex: 1;
    overflow: hidden;
    position: relative;
    cursor: grab;
    background: #fafbfc;
  }
  
  .grid-viewport.dragging {
    cursor: grabbing;
    user-select: none;
  }
  
  .grid-viewport::before {
    content: '';
    position: absolute;
    inset: 0;
    background-image: 
      radial-gradient(circle at 1px 1px, rgba(0,0,0,0.03) 1px, transparent 0);
    background-size: 24px 24px;
    pointer-events: none;
  }
  
  .grid-container {
    display: grid;
    gap: 1px;
    position: absolute;
    top: 0;
    left: 0;
    padding: 20px;
  }
  
  .grid-cell {
    width: 38px;
    height: 38px;
    border-radius: 3px;
    cursor: pointer;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 10px;
    font-weight: 600;
    transition: all 0.15s ease;
    user-select: none;
    position: relative;
  }
  
  .grid-cell::after {
    content: '';
    position: absolute;
    inset: 0;
    border-radius: 3px;
    box-shadow: 0 0 0 1px rgba(0,0,0,0.06);
    pointer-events: none;
  }
  
  .grid-cell:hover {
    transform: scale(1.12);
    z-index: 10;
    box-shadow: 0 4px 12px rgba(0,0,0,0.4);
  }
  
  .cell-normal {
    background: var(--green);
    color: transparent;
  }

  .cell-selecting {
    box-shadow: 0 0 0 2px var(--accent), 0 0 12px rgba(99,102,241,0.4);
    transform: scale(1.12);
    z-index: 10;
  }
  
  .cell-normal:hover { color: white; }
  
  .cell-shield {
    background: var(--red);
    color: transparent;
  }
  
  .cell-shield:hover { color: white; }
  
  .cell-header {
    background: var(--bg-tertiary);
    color: var(--text-secondary);
    font-size: 11px;
    font-weight: 600;
    cursor: default;
  }
  
  .cell-header:hover {
    transform: none;
    box-shadow: none;
  }
  
  .cell-corner {
    background: var(--bg-secondary);
    border: 1px solid var(--border);
  }
  
  /* Sidebar */
  .sidebar {
    width: 300px;
    background: var(--bg-secondary);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    display: flex;
    flex-direction: column;
    overflow: hidden;
    flex-shrink: 0;
  }
  
  .sidebar-header {
    padding: 16px;
    border-bottom: 1px solid var(--border);
  }
  
  .sidebar-title {
    font-size: 14px;
    font-weight: 600;
    color: var(--text-primary);
    display: flex;
    align-items: center;
    gap: 8px;
  }
  
  .sidebar-title img {
    width: 18px;
    height: 18px;
    flex-shrink: 0;
  }
  
  .sidebar-content {
    flex: 1;
    overflow-y: auto;
    padding: 16px;
  }
  
  /* Accordion */
  .accordion {
    margin-bottom: 8px;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    overflow: hidden;
  }
  
  .accordion-header {
    padding: 12px 14px;
    background: var(--bg-tertiary);
    cursor: pointer;
    display: flex;
    justify-content: space-between;
    align-items: center;
    font-size: 13px;
    font-weight: 500;
    color: var(--text-primary);
    transition: var(--transition);
  }
  
  .accordion-header:hover {
    background: var(--bg-card);
  }
  
  .accordion-header img {
    width: 16px;
    height: 16px;
    flex-shrink: 0;
  }
  
  .accordion-header .arrow {
    color: var(--text-muted);
    transition: transform 0.2s;
  }
  
  .accordion-header.open .arrow {
    transform: rotate(180deg);
  }
  
  .accordion-content {
    max-height: 0;
    overflow: hidden;
    transition: max-height 0.25s ease-out;
  }
  
  .accordion-content.open {
    max-height: 400px;
  }
  
  .accordion-inner {
    padding: 14px;
    background: var(--bg-card);
  }
  
  .config-row {
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 8px;
  }
  
  .config-row:last-child {
    margin-bottom: 0;
  }
  
  .config-row label {
    font-size: 12px;
    color: var(--text-secondary);
    min-width: 24px;
    font-weight: 500;
  }
  
  .config-row input {
    flex: 1;
    padding: 7px 10px;
    border-radius: 6px;
    border: 1px solid var(--border);
    background: var(--bg-tertiary);
    color: var(--text-primary);
    font-size: 13px;
    transition: var(--transition);
    outline: none;
  }
  
  .config-row input:focus {
    border-color: var(--accent);
    box-shadow: 0 0 0 2px var(--accent-subtle);
  }
  
  .config-row input::placeholder {
    color: var(--text-muted);
  }
  
  /* Legend */
  .legend {
    display: flex;
    flex-wrap: wrap;
    gap: 12px;
    padding: 14px;
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    margin-bottom: 12px;
  }
  
  .legend-item {
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 12px;
    color: var(--text-secondary);
  }
  
  .legend-dot {
    width: 10px;
    height: 10px;
    border-radius: 2px;
  }
  
  .legend-dot.normal { background: var(--green); }
  .legend-dot.shield { background: var(--red); }
  
  /* Stats */
  .stats {
    padding: 14px;
    background: var(--bg-tertiary);
    border-radius: var(--radius-sm);
    border: 1px solid var(--border);
    margin-bottom: 12px;
  }
  
  .stat-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
    font-size: 12px;
    padding: 4px 0;
    color: var(--text-secondary);
  }
  
  .stat-row span:last-child {
    color: var(--text-primary);
    font-weight: 500;
    font-variant-numeric: tabular-nums;
  }
  
  /* Rule Box */
  .rule-box {
    padding: 12px 14px;
    background: var(--orange-subtle);
    border: 1px solid rgba(245,158,11,0.2);
    border-radius: var(--radius-sm);
    margin-bottom: 12px;
    font-size: 12px;
    color: var(--orange);
    line-height: 1.6;
  }
  
  .rule-box strong {
    color: var(--text-primary);
    font-weight: 600;
  }
  
  /* Export Button */
  .btn-export {
    width: 100%;
    padding: 10px;
    background: var(--accent);
    border: 1px solid var(--accent);
    border-radius: var(--radius-sm);
    color: white;
    font-size: 13px;
    font-weight: 600;
    cursor: pointer;
    transition: var(--transition);
    display: flex;
    align-items: center;
    justify-content: center;
    gap: 6px;
  }
  
  .btn-export img {
    width: 16px;
    height: 16px;
    flex-shrink: 0;
  }
  
  .btn-export:hover {
    background: var(--accent-hover);
    border-color: var(--accent-hover);
  }
  
  /* Drop Zone */
  .drop-zone {
    position: absolute;
    inset: 0;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    background: var(--bg-secondary);
    cursor: pointer;
    transition: var(--transition);
    z-index: 20;
  }
  
  .drop-zone:hover {
    background: var(--bg-tertiary);
  }
  
  .drop-zone-icon {
    width: 64px;
    height: 64px;
    background: var(--accent-subtle);
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    margin-bottom: 16px;
    color: var(--accent);
  }
  
  .drop-zone-text {
    font-size: 15px;
    color: var(--text-primary);
    font-weight: 500;
    margin-bottom: 6px;
  }
  
  .drop-zone-hint {
    font-size: 13px;
    color: var(--text-muted);
  }
  
  /* Version */
  .version {
    position: fixed;
    bottom: 4px;
    right: 16px;
    font-size: 11px;
    color: var(--text-muted);
    font-variant-numeric: tabular-nums;
    pointer-events: none;
    z-index: 30;
  }
  
  /* Toast */
  .toast {
    position: fixed;
    top: 72px;
    left: 50%;
    transform: translateX(-50%) translateY(-20px);
    padding: 10px 20px;
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    box-shadow: var(--shadow);
    font-size: 13px;
    color: var(--text-primary);
    opacity: 0;
    pointer-events: none;
    transition: all 0.3s ease;
    z-index: 1000;
  }
  
  .toast.show {
    opacity: 1;
    transform: translateX(-50%) translateY(0);
  }
  
  /* Mobile */
  @media (max-width: 768px) {
    .main { flex-direction: column; height: auto; }
    .sidebar { width: 100%; }
    .header { padding: 12px 16px; }
    .logo-text { display: none; }
  }
</style>
</head>
<body>

<header class="header">
  <a class="logo" href="#">
    <div class="logo-icon">
      <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z"/><polyline points="3.27 6.96 12 12.01 20.73 6.96"/><line x1="12" y1="22.08" x2="12" y2="12"/></svg>
    </div>
    <span class="logo-text">库位编辑器</span>
  </a>
  <div class="header-actions">
    <button class="btn" onclick="document.getElementById('fileInput').click()">
      <img src="/icons/upload.svg" alt="上传">
      <span>导入</span>
    </button>
    <input type="file" id="fileInput" accept=".xlsx,.xls" style="display:none" onchange="handleFile(this.files[0])">
    <button class="btn" onclick="resetShields()">
      <img src="/icons/rotate.svg" alt="重置">
      <span>重置</span>
    </button>
  </div>
</header>

<div class="main">
  <div class="grid-area">
    <div class="drop-zone" id="dropZone" ondrop="dropHandler(event)" ondragover="dragOver(event)" ondragleave="dragLeave(event)" onclick="document.getElementById('fileInput').click()">
      <div class="drop-zone-icon">
        <svg xmlns="http://www.w3.org/2000/svg" width="28" height="28" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
      </div>
      <div class="drop-zone-text">点击或拖拽导入 Excel</div>
      <div class="drop-zone-hint">支持 .xlsx 格式</div>
    </div>
    
    <div class="grid-toolbar" id="gridToolbar" style="display:none" onclick="if(ctrlLocked){detailLocked=false;ctrlLocked=false}">
      <div class="z-tabs" id="zTabs"></div>
      <div class="toolbar-spacer"></div>
      <button id="btnSyncLayers" class="btn btn-sm" onclick="toggleSyncLayers()" title="同步所有层：修改Z1时所有层一起变">
      <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><polyline points="17 1 21 5 17 9"/><path d="M3 11V9a4 4 0 0 1 4-4h14"/><polyline points="7 23 3 19 7 15"/><path d="M21 13v2a4 4 0 0 1-4 4H3"/></svg>
      <span>同步层</span>
    </button>
      <div class="zoom-controls">
        <button class="zoom-btn" onclick="zoomOut()"><img src="/icons/minus.svg" alt="缩小"></button>
        <span class="zoom-level" id="zoomLevel">100%</span>
        <button class="zoom-btn" onclick="zoomIn()"><img src="/icons/plus.svg" alt="放大"></button>
        <button class="zoom-btn" onclick="toggleXYSwap()" title="交换XY轴"><span style="font-size:12px;font-weight:600;">XY</span></button>
        <button class="zoom-btn" onclick="flipXAxis()" title="翻转X轴"><span style="font-size:12px;font-weight:600;">X↔</span></button>
        <button class="zoom-btn" onclick="flipYAxis()" title="翻转Y轴"><span style="font-size:12px;font-weight:600;">Y↕</span></button>
      </div>
    </div>
    
    <div class="grid-viewport" id="gridViewport" style="display:none">
      <div class="grid-container" id="gridContainer"></div>
    </div>
  </div>
  
  <div class="sidebar" id="sidebar" style="display:none">
    <div class="sidebar-header">
      <div class="sidebar-title">
        <img src="/icons/box.svg" alt="统计">
        参数设置
      </div>
    </div>
    
    <div class="sidebar-content">
      <div class="accordion">
        <div class="accordion-header open" onclick="toggleAccordion(this)">
          <span style="display:flex;align-items:center;gap:6px"><img src="/icons/ruler.svg" alt="库位详情">库位详情</span>
          <span class="arrow"><img src="/icons/chevron.svg" alt="展开"></span>
        </div>
        <div class="accordion-content open" id="cellDetailPanel">
          <div class="accordion-inner" id="cellDetailInner">
            <div style="color:var(--text-muted);font-size:12px;text-align:center;padding:20px 0;">点击格子查看详情</div>
          </div>
        </div>
      </div>
      
      <div class="legend">
        <div class="legend-item"><div class="legend-dot normal"></div>正常</div>
        <div class="legend-item"><div class="legend-dot shield"></div>不可用</div>
      </div>
      
      <div class="stats" id="stats"></div>
      
      <button class="btn-export" onclick="exportFile()">
        <img src="/icons/download.svg" alt="导出">
        导出 Excel
      </button>
    </div>
  </div>
</div>

<div class="version">v6.8</div>
<div class="toast" id="toast"></div>

<script>
// 会话 ID，隔离多用户数据
let sessionId = localStorage.getItem('wh_session_id');
if (!sessionId) {
  sessionId = crypto.randomUUID();
  localStorage.setItem('wh_session_id', sessionId);
}

let appData = null;
let originalFileName = 'warehouse';
let layerConfigs = {}; 
let zoom = 1;
let currentZ = 1;
let isDragging = false;
let dragStart = { x: 0, y: 0 };
let gridOffset = { x: 0, y: 0 };

// 批量选择状态
let isSelecting = false;
let selectionStart = { x: 0, y: 0 };
let ctrlHeld = false; // Ctrl+点击不修改状态
let selectionEnd = { x: 0, y: 0 };
let selectedCells = new Set();
let batchDialogVisible = false;

// 按层存储禁用状态：{ z: Set("x,y") }
let manualShieldedByLayer = {};
let syncLayersEnabled = false;
let selectedCellXY = null; // 当前选中的格子 {x, y}
let detailLocked = false; // 详情面板锁定（点击格子后锁定，悬停不切换）
let ctrlLocked = false; // Ctrl+点击锁定（点击toolbar可解锁）
let cellConfigs = {}; // 每个格子独立的层高/承重: { "x,y,z": { height, weight } }

function getShieldedLayer(z) {
  if (!manualShieldedByLayer[z]) manualShieldedByLayer[z] = new Set();
  return manualShieldedByLayer[z];
}

function isCellShielded(x, y, z) {
  return getShieldedLayer(z).has(`${x},${y}`);
}

function dragOver(e) { e.preventDefault(); }
function dropHandler(e) { e.preventDefault(); handleFile(e.dataTransfer.files[0]); }

function toggleAccordion(el) {
  el.classList.toggle('open');
  el.nextElementSibling.classList.toggle('open');
}

function showToast(msg) {
  const toast = document.getElementById('toast');
  toast.textContent = msg;
  toast.classList.add('show');
  setTimeout(() => toast.classList.remove('show'), 2000);
}

async function handleFile(file) {
  if (!file) return;
  appData = null;
  originalFileName = file.name.replace(/\.xlsx$/, '');
  const fd = new FormData();
  fd.append('file', file);
  const resp = await fetch('/upload', { method: 'POST', body: fd, headers: { 'X-Session-Id': sessionId } });
  const data = await resp.json();
  if (data.error) { showToast(data.error); return; }
  
  appData = data;
  manualShieldedByLayer = {};
  // 按层分发屏蔽状态
  (data.shielded || []).forEach(s => {
    const parts = s.split(',');
    const z = parts.length >= 3 ? parseInt(parts[2]) : 1;
    const key = parts[0] + ',' + parts[1];
    if (!manualShieldedByLayer[z]) manualShieldedByLayer[z] = new Set();
    manualShieldedByLayer[z].add(key);
  });
  layerConfigs = data.layer_configs || {};
  // 恢复之前保存的 cellConfigs
  if (data.saved_cell_configs && Object.keys(data.saved_cell_configs).length > 0) {
    cellConfigs = data.saved_cell_configs;
    showToast(`已恢复 ${Object.keys(cellConfigs).length} 个已保存的配置`);
  }
  zoom = 1;
  gridOffset = { x: 0, y: 0 };
  
  // 确保所有层都有初始化
  appData.z_range.forEach(z => {
    if (!manualShieldedByLayer[z]) manualShieldedByLayer[z] = new Set();
  });
  
  // X=竖向(行)从上到下, Y=横向(列)从左到右，都从左上角开始
  appData.x_range.sort((a, b) => a - b);
  appData.y_range.sort((a, b) => a - b);
  
  document.getElementById('dropZone').style.display = 'none';
  document.getElementById('gridToolbar').style.display = 'flex';
  document.getElementById('gridViewport').style.display = 'block';
  document.getElementById('sidebar').style.display = 'flex';
  
  renderZTabs();
  renderGrid();
  fitToViewport();
  updateStats();
}

function renderZTabs() {
  const tabs = document.getElementById('zTabs');
  tabs.innerHTML = '';
  appData.z_range.forEach(z => {
    const tab = document.createElement('div');
    tab.className = 'z-tab' + (z === currentZ ? ' active' : '');
    tab.textContent = 'Z=' + z;
    tab.onclick = () => {
      currentZ = z;
      detailLocked = false; // 切换层时解锁详情面板
      renderZTabs();
      renderGrid();
      updateStats();
    };
    tabs.appendChild(tab);
  });
}

function showCellDetail(x, y, skipStatusToggle) {
  const inner = document.getElementById('cellDetailInner');
  if (!inner) return;
  selectedCellXY = {x, y};
  const key = `${x},${y},${currentZ}`;
  const isShielded = isCellShielded(x, y, currentZ);
  const statusText = isShielded ? '不可用' : '正常';
  const statusColor = isShielded ? 'var(--red)' : 'var(--green)';
  const cfg = cellConfigs[key] || {};
  const layerCfg = layerConfigs[currentZ] || {};
  const h = cfg.height !== undefined ? cfg.height : (layerCfg.height || '');
  const w = cfg.weight !== undefined ? cfg.weight : (layerCfg.weight || '');
  inner.innerHTML = `
    <div style="margin-bottom:12px;">
      <div class="config-row"><label style="min-width:24px;font-size:12px;color:var(--text-secondary);">位置</label><span style="font-size:13px;font-weight:500;">X${x} Y${y} Z${currentZ}</span></div>
      <div class="config-row"><label style="min-width:24px;font-size:12px;color:var(--text-secondary);">状态</label><span style="font-size:13px;font-weight:500;color:${statusColor};">${statusText}</span></div>
    </div>
    <div style="margin-bottom:12px;">
      <div style="font-size:13px;font-weight:600;color:var(--text-primary);margin-bottom:8px;">层高 / 承重</div>
      <div class="config-row">
        <label style="min-width:24px;font-size:12px;color:var(--text-secondary);">层高</label>
        <input type="number" step="0.1" placeholder="m" value="${h}" onchange="updateCellConfig(${x},${y},${currentZ},'height',this.value)">
        <span style="font-size:12px;color:var(--text-muted);">m</span>
      </div>
      <div class="config-row">
        <label style="min-width:24px;font-size:12px;color:var(--text-secondary);">承重</label>
        <input type="number" step="1" placeholder="kg" value="${w}" onchange="updateCellConfig(${x},${y},${currentZ},'weight',this.value)">
        <span style="font-size:12px;color:var(--text-muted);">kg</span>
      </div>
    </div>
    <div style="font-size:11px;color:var(--text-muted);text-align:center;">点击格子可切换状态，<br>ctrl + 点击格子不切换状态，<br>拖选多个可批量操作</div>
  `;
  // 高亮选中格子
  const container = document.getElementById('gridContainer');
  container.querySelectorAll('.grid-cell.selected-cell').forEach(c => c.classList.remove('selected-cell'));
  const target = container.querySelector(`.grid-cell[data-x="${x}"][data-y="${y}"]`);
  if (target) target.classList.add('selected-cell');
}

function updateCellConfig(x, y, z, field, val) {
  const parsedVal = parseFloat(val) || 0;
  if (syncLayersEnabled && appData.z_range.length > 1) {
    // 同步层模式：所有层都改
    appData.z_range.forEach(zz => {
      const key = `${x},${y},${zz}`;
      if (!cellConfigs[key]) cellConfigs[key] = {};
      cellConfigs[key][field] = parsedVal;
    });
    showToast(`已同步所有层保存 X${x} Y${y} ${field === 'height' ? '层高' : '承重'}`);
  } else {
    const key = `${x},${y},${z}`;
    if (!cellConfigs[key]) cellConfigs[key] = {};
    cellConfigs[key][field] = parsedVal;
    showToast(`已保存 X${x} Y${y} Z${z} ${field === 'height' ? '层高' : '承重'}`);
  }
  // 自动保存到服务器
  autoSaveConfig();
}

let saveTimer = null;
function autoSaveConfig() {
  if (saveTimer) clearTimeout(saveTimer);
  saveTimer = setTimeout(async () => {
    try {
      const resp = await fetch('/save-config', {
        method: 'POST',
        headers: {'Content-Type': 'application/json', 'X-Session-Id': sessionId},
        body: JSON.stringify({ cell_configs: cellConfigs })
      });
      const result = await resp.json();
      if (result.ok) {
        console.log('配置已自动保存');
      }
    } catch (e) {
      console.error('自动保存失败:', e);
    }
  }, 500); // 500ms 防抖
}

function renderGrid() {
  const container = document.getElementById('gridContainer');
  container.innerHTML = '';
  
  const xs = appData.x_range;
  const ys = appData.y_range;
  
  // 根据 xySwapped 决定哪个是行、哪个是列
  const rowItems = xySwapped ? ys : xs;
  const colItems = xySwapped ? xs : ys;
  const rowLabel = xySwapped ? 'Y' : 'X';
  const colLabel = xySwapped ? 'X' : 'Y';
  
  const cols = colItems.length + 1;
  container.style.gridTemplateColumns = `repeat(${cols}, 38px)`;
  
  // 左上角空白
  const corner = document.createElement('div');
  corner.className = 'grid-cell cell-corner';
  container.appendChild(corner);
  
  // 列标题
  colItems.forEach(c => {
    const h = document.createElement('div');
    h.className = 'grid-cell cell-header';
    h.textContent = colLabel + c;
    container.appendChild(h);
  });
  
  // 行
  rowItems.forEach(r => {
    const rh = document.createElement('div');
    rh.className = 'grid-cell cell-header';
    rh.textContent = rowLabel + r;
    container.appendChild(rh);
    
    // 列
    colItems.forEach(c => {
      const x = xySwapped ? c : r;
      const y = xySwapped ? r : c;
      
      const cell = document.createElement('div');
      cell.className = 'grid-cell';
      cell.dataset.x = x;
      cell.dataset.y = y;
      
      const isShielded = isCellShielded(x, y, currentZ);
      
      if (isShielded) {
        cell.classList.add('cell-shield');
      } else {
        cell.classList.add('cell-normal');
      }
      
      let statusText = isShielded ? '不可用' : '正常';
      
      cell.title = `X${x} Y${y} Z${currentZ} (${statusText})`;
      // 所有格子都支持鼠标拖选（包括上层）
      cell.onmousedown = (e) => {
        if (e.button === 2) return; // 右键不触发选中
        e.preventDefault();
        onCellMouseDown(x, y, e);
      };
      // 右键取消锁定 + 清除选中框
      cell.oncontextmenu = (e) => {
        e.preventDefault();
        if (detailLocked) {
          detailLocked = false;
          ctrlLocked = false;
          selectedCellXY = null;
          // 移除选中高亮
          document.querySelectorAll('.grid-cell.selected-cell').forEach(c => c.classList.remove('selected-cell'));
          // 清除选中框（cell-selecting）
          selectedCells.clear();
          updateSelectionHighlight();
          isSelecting = false;
          // 恢复详情面板默认提示
          document.getElementById('cellDetailInner').innerHTML = '<div style="color:var(--text-muted);font-size:12px;text-align:center;padding:20px 0;">点击格子查看详情</div>';
          console.log('右键取消锁定+清除选中框');
        }
      };
      cell.onmouseover = () => {
        if (isSelecting) onCellMouseOver(x, y);
        else if (!detailLocked) showCellDetail(x, y);
        else console.log('detailLocked, skip hover');
      };
      cell.onmouseup = () => {
        if (isSelecting) onCellMouseUp();
      };
      container.appendChild(cell);
    });
  });
  
  applyZoom();
}

// 批量选择逻辑
function onCellMouseDown(x, y, e) {
  if (batchDialogVisible) return;
  isSelecting = true;
  selectionStart = { x, y };
  selectionEnd = { x, y };
  selectedCells = new Set([`${x},${y}`]);
  ctrlHeld = e.ctrlKey || e.metaKey;
  updateSelectionHighlight();
}

function onCellMouseOver(x, y) {
  if (!isSelecting) return;
  selectionEnd = { x, y };
  selectedCells = getCellInRange(selectionStart, selectionEnd);
  updateSelectionHighlight();
}

function onCellMouseUp() {
  if (!isSelecting) return;
  isSelecting = false;
  if (selectedCells.size === 1) {
    const cell = [...selectedCells][0];
    const [x, y] = cell.split(',').map(Number);
    // 单个：Ctrl+点击不切换状态
    if (!ctrlHeld) {
      toggleCell(x, y, currentZ);
    }
    // toggle之后再显示详情，确保状态和格子颜色一致
    showCellDetail(x, y);
    detailLocked = true; // 锁定详情面板，悬停不再切换
    ctrlLocked = ctrlHeld; // Ctrl+点击时标记
    console.log('Cell clicked:', {x, y, ctrlHeld, detailLocked, ctrlLocked});
  } else if (selectedCells.size > 1) {
    // 多个：弹批量确认框
    showBatchDialog();
  }
}

function getCellInRange(start, end) {
  const minX = Math.min(start.x, end.x);
  const maxX = Math.max(start.x, end.x);
  const minY = Math.min(start.y, end.y);
  const maxY = Math.max(start.y, end.y);
  
  const cells = new Set();
  for (let x = minX; x <= maxX; x++) {
    for (let y = minY; y <= maxY; y++) {
      // 所有格子都可以进入批量选择，包括上层
      cells.add(`${x},${y}`);
    }
  }
  return cells;
}

function updateSelectionHighlight() {
  const container = document.getElementById('gridContainer');
  const cells = container.querySelectorAll('.grid-cell[data-x]');
  cells.forEach(cell => {
    const x = parseInt(cell.dataset.x);
    const y = parseInt(cell.dataset.y);
    if (selectedCells.has(`${x},${y}`)) {
      cell.classList.add('cell-selecting');
    } else {
      cell.classList.remove('cell-selecting');
    }
  });
}

function showBatchDialog() {
  batchDialogVisible = true;
  const minX = Math.min(selectionStart.x, selectionEnd.x);
  const maxX = Math.max(selectionStart.x, selectionEnd.x);
  const minY = Math.min(selectionStart.y, selectionEnd.y);
  const maxY = Math.max(selectionStart.y, selectionEnd.y);
  
  // 统计当前层选中格子的状态分布
  let curEnabled = 0;  // 当前启用（正常）的
  let curDisabled = 0; // 当前禁用（不可用）的
  const layer = getShieldedLayer(currentZ);
  selectedCells.forEach(key => {
    if (layer.has(key)) curDisabled++;
    else curEnabled++;
  });
  
  const hasMixed = curEnabled > 0 && curDisabled > 0;
  
  const dialog = document.createElement('div');
  dialog.id = 'batchDialog';
  dialog.style.cssText = `
    position: fixed;
    top: 50%;
    left: 50%;
    transform: translate(-50%, -50%);
    background: white;
    border-radius: 12px;
    padding: 24px;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
    z-index: 2000;
    min-width: 280px;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
  `;
  
  if (hasMixed) {
    // 混合状态：两个按钮，标注各自会改多少个
    dialog.innerHTML = `
      <div style="font-size:16px;font-weight:600;margin-bottom:12px;color:#1a1a2e;">
        批量操作
      </div>
      <div style="font-size:13px;color:#5c5c70;margin-bottom:16px;">
        范围：X${minX}-${maxX}，Y${minY}-${maxY}，Z${currentZ}<br>
        已选 <strong>${selectedCells.size}</strong> 个库位
        （当前启用 ${curEnabled}，当前禁用 ${curDisabled}）
      </div>
      <div style="margin-bottom:12px;">
        <div class="config-row"><label style="min-width:24px;font-size:12px;color:#5c5c70;">层高</label><input id="batchHeight" type="number" step="0.1" placeholder="m" style="flex:1;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:13px;"><span style="font-size:12px;color:#999;">m</span></div>
        <div class="config-row"><label style="min-width:24px;font-size:12px;color:#5c5c70;">承重</label><input id="batchWeight" type="number" step="1" placeholder="kg" style="flex:1;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:13px;"><span style="font-size:12px;color:#999;">kg</span></div>
      </div>
      <div style="display:flex;gap:8px;">
        <button id="btnBatchEnable" style="flex:1;padding:10px;border:none;border-radius:8px;background:#22c55e;color:white;font-size:13px;font-weight:600;cursor:pointer;">
          批量启用 (${curDisabled}个)
        </button>
        <button id="btnBatchDisable" style="flex:1;padding:10px;border:none;border-radius:8px;background:#ef4444;color:white;font-size:13px;font-weight:600;cursor:pointer;">
          批量禁用 (${curEnabled}个)
        </button>
      </div>
      <button id="btnBatchOnlyConfig" style="width:100%;margin-top:8px;padding:8px;border:1px solid #ddd;border-radius:8px;background:white;color:#1a1a2e;font-size:13px;cursor:pointer;">
        仅修改层高/承重（不改状态）
      </button>
      <button id="btnBatchCancel" style="width:100%;margin-top:8px;padding:8px;border:1px solid #ddd;border-radius:8px;background:white;color:#5c5c70;font-size:13px;cursor:pointer;">
        取消
      </button>
    `;
  } else {
    // 全是启用 → 只能禁用；全是禁用 → 只能启用
    const willDisable = curEnabled > 0;
    const actionLabel = willDisable ? '禁用' : '启用';
    const actionColor = willDisable ? '#ef4444' : '#22c55e';
    const actionCount = willDisable ? curEnabled : curDisabled;
    
    dialog.innerHTML = `
      <div style="font-size:16px;font-weight:600;margin-bottom:12px;color:#1a1a2e;">
        批量操作
      </div>
      <div style="font-size:13px;color:#5c5c70;margin-bottom:16px;">
        范围：X${minX}-${maxX}，Y${minY}-${maxY}，Z${currentZ}<br>
        已选 <strong>${selectedCells.size}</strong> 个库位，全部将被${actionLabel}
      </div>
      <div style="margin-bottom:12px;">
        <div class="config-row"><label style="min-width:24px;font-size:12px;color:#5c5c70;">层高</label><input id="batchHeight" type="number" step="0.1" placeholder="m" style="flex:1;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:13px;"><span style="font-size:12px;color:#999;">m</span></div>
        <div class="config-row"><label style="min-width:24px;font-size:12px;color:#5c5c70;">承重</label><input id="batchWeight" type="number" step="1" placeholder="kg" style="flex:1;padding:6px 8px;border:1px solid #ddd;border-radius:6px;font-size:13px;"><span style="font-size:12px;color:#999;">kg</span></div>
      </div>
      <div style="display:flex;gap:8px;">
        <button id="btnBatchConfirm" style="flex:1;padding:10px;border:none;border-radius:8px;background:${actionColor};color:white;font-size:13px;font-weight:600;cursor:pointer;">
          批量${actionLabel} (${actionCount}个)
        </button>
        <button id="btnBatchCancel" style="flex:1;padding:10px;border:1px solid #ddd;border-radius:8px;background:white;color:#5c5c70;font-size:13px;cursor:pointer;">
          取消
        </button>
      </div>
      <button id="btnBatchOnlyConfig" style="width:100%;margin-top:8px;padding:8px;border:1px solid #ddd;border-radius:8px;background:white;color:#1a1a2e;font-size:13px;cursor:pointer;">
        仅修改层高/承重（不改状态）
      </button>
    `;
  }
  
  document.body.appendChild(dialog);
  
  if (hasMixed) {
    document.getElementById('btnBatchEnable').onclick = () => batchToggleStatus('enabled');
    document.getElementById('btnBatchDisable').onclick = () => batchToggleStatus('disabled');
  } else {
    const willDisable = curEnabled > 0;
    document.getElementById('btnBatchConfirm').onclick = () => batchToggleStatus(willDisable ? 'disabled' : 'enabled');
  }
  document.getElementById('btnBatchOnlyConfig').onclick = () => batchToggleStatus(null);
  document.getElementById('btnBatchCancel').onclick = closeBatchDialog;
}

function closeBatchDialog() {
  batchDialogVisible = false;
  selectedCells.clear();
  const dialog = document.getElementById('batchDialog');
  if (dialog) dialog.remove();
  updateSelectionHighlight();
}

function batchToggleStatus(status) {
  let changedCount = 0;
  const batchH = document.getElementById('batchHeight');
  const batchW = document.getElementById('batchWeight');
  const hVal = batchH && batchH.value !== '' ? parseFloat(batchH.value) : undefined;
  const wVal = batchW && batchW.value !== '' ? parseFloat(batchW.value) : undefined;
  const zsToModify = syncLayersEnabled ? appData.z_range : [currentZ];
  selectedCells.forEach(key => {
    zsToModify.forEach(z => {
      const layer = getShieldedLayer(z);
      const isCurrentlyShielded = layer.has(key);
      if (status === 'disabled' && !isCurrentlyShielded) {
        layer.add(key);
        changedCount++;
      } else if (status === 'enabled' && isCurrentlyShielded) {
        layer.delete(key);
        changedCount++;
      }
      // 批量设置层高/承重
      const [x, y] = key.split(',').map(Number);
      const cellKey = `${x},${y},${z}`;
      if (!cellConfigs[cellKey]) cellConfigs[cellKey] = {};
      if (hVal !== undefined) cellConfigs[cellKey].height = hVal;
      if (wVal !== undefined) cellConfigs[cellKey].weight = wVal;
    });
  });
  closeBatchDialog();
  renderGrid();
  updateStats();
  if (changedCount === 0 && hVal === undefined && wVal === undefined) {
    showToast(`没有需要${status === 'enabled' ? '启用' : '禁用'}的库位`);
  } else {
    let msg = '';
    if (changedCount > 0) msg += `已${status === 'enabled' ? '启用' : '禁用'} ${changedCount} 个库位`;
    if (hVal !== undefined || wVal !== undefined) {
      const count = selectedCells.size * zsToModify.length;
      const parts = [];
      if (hVal !== undefined) parts.push(`层高 ${hVal}m`);
      if (wVal !== undefined) parts.push(`承重 ${wVal}kg`);
      msg += (msg ? '，' : '') + `已设置 ${count} 个库位：${parts.join('，')}`;
    }
    showToast(msg);
  }
}

function toggleSyncLayers() {
  syncLayersEnabled = !syncLayersEnabled;
  const btn = document.getElementById('btnSyncLayers');
  if (syncLayersEnabled) {
    btn.classList.add('btn-primary');
    btn.title = '同步层已开启：修改Z1时所有层一起变';
  } else {
    btn.classList.remove('btn-primary');
    btn.title = '同步所有层：修改Z1时所有层一起变';
  }
}

function toggleCell(x, y, z) {
  const key = `${x},${y}`;
  const layer = getShieldedLayer(z);
  const isCurrentlyShielded = layer.has(key);
  
  // 同步层模式：同时修改所有层
  if (syncLayersEnabled && appData.z_range.length > 1) {
    appData.z_range.forEach(zz => {
      const l = getShieldedLayer(zz);
      if (isCurrentlyShielded) {
        l.delete(key);
      } else {
        l.add(key);
      }
    });
  } else {
    if (isCurrentlyShielded) {
      layer.delete(key);
    } else {
      layer.add(key);
    }
  }
  renderGrid();
  updateStats();
}

function resetShields() {
  if (!appData) return;
  manualShieldedByLayer = {};
  appData.z_range.forEach(z => {
    manualShieldedByLayer[z] = new Set();
  });
  renderGrid();
  updateStats();
  showToast('已重置');
}

function updateStats() {
  const total = appData.x_range.length * appData.y_range.length;
  const layer = getShieldedLayer(currentZ);
  const curShield = layer.size;
  
  // 总库位（正常）= 所有楼层的可用状态总数
  let totalNormal = 0;
  appData.z_range.forEach(z => {
    const l = getShieldedLayer(z);
    totalNormal += total - l.size;
  });
  
  document.getElementById('stats').innerHTML = `
    <div class="stat-row"><span>X 轴</span><span>${appData.x_range.length}</span></div>
    <div class="stat-row"><span>Y 轴</span><span>${appData.y_range.length}</span></div>
    <div class="stat-row"><span>Z 层数</span><span>${appData.z_range.length}</span></div>
    <div class="stat-row"><span>总库位（正常）</span><span>${totalNormal}</span></div>
    <div class="stat-row"><span>当前层禁用</span><span>${curShield}</span></div>
    <div class="stat-row"><span>当前层可用</span><span>${total - curShield}</span></div>
  `;
}

function fitToViewport() {
  const viewport = document.getElementById('gridViewport');
  const container = document.getElementById('gridContainer');
  
  if (!viewport || !container || !appData) return;
  
  const vpWidth = viewport.clientWidth - 40;
  const vpHeight = viewport.clientHeight - 40;
  const xs = appData.x_range.length;
  const ys = appData.y_range.length;
  const cellSize = 38;
  const gap = 1;
  // 根据 xySwapped 决定宽高计算
  const colCount = xySwapped ? xs : ys;
  const rowCount = xySwapped ? ys : xs;
  const totalWidth = (colCount + 1) * (cellSize + gap);
  const totalHeight = (rowCount + 1) * (cellSize + gap);
  
  zoom = Math.min(vpWidth / totalWidth, vpHeight / totalHeight, 1);
  gridOffset = {
    x: (vpWidth - totalWidth * zoom) / 2,
    y: (vpHeight - totalHeight * zoom) / 2
  };
  
  container.style.transformOrigin = 'top left';
  applyZoom();
}

function zoomIn() {
  zoom = Math.min(zoom * 1.2, 3);
  applyZoom();
}

function zoomOut() {
  zoom = Math.max(zoom / 1.2, 0.3);
  applyZoom();
}

function applyZoom() {
  const container = document.getElementById('gridContainer');
  container.style.transform = `scale(${zoom}) translate(${gridOffset.x / zoom}px, ${gridOffset.y / zoom}px)`;
  document.getElementById('zoomLevel').textContent = Math.round(zoom * 100) + '%';
}

let xySwapped = false;
function toggleXYSwap() {
  xySwapped = !xySwapped;
  renderGrid();
  fitToViewport();
}
function flipXAxis() {
  appData.x_range.reverse();
  renderGrid();
  updateStats();
  showToast('X轴已翻转');
}
function flipYAxis() {
  appData.y_range.reverse();
  renderGrid();
  updateStats();
  showToast('Y轴已翻转');
}

const viewport = document.getElementById('gridViewport');
viewport.addEventListener('mousedown', (e) => {
  if (e.target.classList.contains('grid-cell') && !e.target.classList.contains('cell-header')) return;
  isDragging = true;
  dragStart = { x: e.clientX - gridOffset.x, y: e.clientY - gridOffset.y };
  viewport.classList.add('dragging');
});
// 右键点击取消锁定 + 清除选中框
viewport.addEventListener('contextmenu', (e) => {
  e.preventDefault();
  if (detailLocked) {
    detailLocked = false;
    ctrlLocked = false;
    selectedCells.clear();
    updateSelectionHighlight();
    isSelecting = false;
    console.log('右键取消锁定+清除选中框');
  }
});

document.addEventListener('mousemove', (e) => {
  if (!isDragging) return;
  gridOffset.x = e.clientX - dragStart.x;
  gridOffset.y = e.clientY - dragStart.y;
  applyZoom();
});

document.addEventListener('mouseup', () => {
  isDragging = false;
  viewport.classList.remove('dragging');
});

viewport.addEventListener('touchstart', (e) => {
  if (e.target.classList.contains('grid-cell') && !e.target.classList.contains('cell-header')) return;
  if (e.touches.length === 1) {
    isDragging = true;
    dragStart = { x: e.touches[0].clientX - gridOffset.x, y: e.touches[0].clientY - gridOffset.y };
  }
});

viewport.addEventListener('touchmove', (e) => {
  if (!isDragging || e.touches.length !== 1) return;
  e.preventDefault();
  gridOffset.x = e.touches[0].clientX - dragStart.x;
  gridOffset.y = e.touches[0].clientY - dragStart.y;
  applyZoom();
}, { passive: false });

viewport.addEventListener('touchend', () => { isDragging = false; });

async function exportFile() {
  if (!appData) return;
  
  // 导出：每层独立禁用状态
  const allShielded = [];
  appData.z_range.forEach(z => {
    const layer = getShieldedLayer(z);
    layer.forEach(key => {
      const [x, y] = key.split(',');
      allShielded.push(`${x},${y},${z}`);
    });
  });
  
  const payload = JSON.stringify({
    shielded: allShielded,
    layer_configs: layerConfigs,
    cell_configs: cellConfigs,
    x_range: appData.x_range,
    y_range: appData.y_range,
    z_range: appData.z_range,
    warehouse_name: appData.warehouse_name || '',
    zone_name: appData.zone_name || '',
    cell_data: appData.cell_data || {}
  });
  
  const resp = await fetch('/export', { method: 'POST', headers: {'Content-Type':'application/json', 'X-Session-Id': sessionId}, body: payload });
  const blob = await resp.blob();
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = originalFileName + '_edited.xlsx';
  a.click();
  showToast('导出成功！');
}
</script>
</body>
</html>"""

# ============= Request Handler =============

class Handler(http.server.BaseHTTPRequestHandler):
    _sessions = {}  # { session_id: { 'bytes': wb_bytes, 'ts': timestamp } }
    _lock = threading.Lock()
    
    def _get_session_id(self):
        return self.headers.get('X-Session-Id', 'default')
    
    def _cleanup_sessions(self):
        """清理超过1小时未使用的会话"""
        now = time.time()
        expired = [k for k, v in self._sessions.items() if now - v['ts'] > 3600]
        for k in expired:
            del self._sessions[k]

    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(HTML_PAGE.encode('utf-8'))
        elif self.path.startswith('/icons/'):
            icon_name = self.path.split('/')[-1]
            icon_path = f'/root/icons/{icon_name}'
            if os.path.exists(icon_path):
                self.send_response(200)
                self.send_header('Content-Type', 'image/svg+xml')
                self.end_headers()
                with open(icon_path, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.send_error(404)
        else:
            self.send_error(404)
    
    def do_POST(self):
        if self.path == '/upload':
            self.handle_upload()
        elif self.path == '/export':
            self.handle_export()
        elif self.path == '/save-config':
            self.handle_save_config()
        else:
            self.send_error(404)
    
    def handle_upload(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        
        boundary = self.headers.get('Content-Type', '').split('boundary=')[-1]
        if not boundary:
            self.json_response({'error': 'Invalid upload'})
            return
        
        parts = body.split(("--" + boundary).encode())
        file_data = None
        for part in parts:
            if b'Content-Disposition' in part and b'filename=' in part:
                idx = part.find(b'\r\n\r\n')
                if idx > 0:
                    file_data = part[idx+4:]
                    if file_data.endswith(b'\r\n'):
                        file_data = file_data[:-2]
                break
        
        if not file_data:
            self.json_response({'error': 'No file data'})
            return
        
        try:
            # 保存原始工作簿（不加 data_only，保留公式和格式）
            wb_orig = openpyxl.load_workbook(io.BytesIO(file_data))
            buf = io.BytesIO()
            wb_orig.save(buf)
            session_id = self._get_session_id()
            with Handler._lock:
                # 如果会话已存在，保留已有的 cell_configs
                existing_configs = {}
                if session_id in Handler._sessions:
                    existing_configs = Handler._sessions[session_id].get('cell_configs', {})
                Handler._sessions[session_id] = {
                    'bytes': buf.getvalue(),
                    'ts': time.time(),
                    'cell_configs': existing_configs  # 保留之前保存的配置
                }
                self._cleanup_sessions()
            
            # 解析用 data_only 版本（需要读取公式的计算结果）
            result = parse_excel(file_data)
            # 附加已保存的 cell_configs
            result['saved_cell_configs'] = Handler._sessions[session_id].get('cell_configs', {})
            self.json_response(result)
        except Exception as e:
            self.json_response({'error': str(e)})
    
    def handle_save_config(self):
        """保存单个格子的层高/承重配置"""
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        try:
            data = json.loads(body)
            session_id = self._get_session_id()
            
            with Handler._lock:
                if session_id in Handler._sessions:
                    # 保存配置到会话
                    if 'cell_configs' not in Handler._sessions[session_id]:
                        Handler._sessions[session_id]['cell_configs'] = {}
                    
                    # 合并新的配置
                    new_configs = data.get('cell_configs', {})
                    Handler._sessions[session_id]['cell_configs'].update(new_configs)
                    
                    self.json_response({'ok': True, 'saved': len(new_configs)})
                else:
                    self.json_response({'error': 'Session not found'})
        except Exception as e:
            self.json_response({'error': str(e)})
    
    def handle_export(self):
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        try:
            data = json.loads(body)
            shielded = data.get('shielded', [])
            layer_configs = data.get('layer_configs', {})
            cell_configs = data.get('cell_configs', {})
            x_range = data.get('x_range', [])
            y_range = data.get('y_range', [])
            z_range = data.get('z_range', [])
            cell_data = data.get('cell_data', {})
            warehouse_name = data.get('warehouse_name', '')
            zone_name = data.get('zone_name', '')
            
            shielded_set = set()
            for s in shielded:
                parts = s.split(',')
                px, py = int(parts[0]), int(parts[1])
                pz = int(parts[2]) if len(parts) >= 3 else 1
                shielded_set.add((px, py, pz))
            
            cfg = {int(k): v for k, v in layer_configs.items()}
            
            # 用 generate_excel 生成新 Excel
            excel_bytes = generate_excel(x_range, y_range, z_range, shielded_set, cfg, cell_data, warehouse_name, zone_name, cell_configs)
            
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="warehouse_export.xlsx"')
            self.send_header('Content-Length', str(len(excel_bytes)))
            self.end_headers()
            self.wfile.write(excel_bytes)
        except Exception as e:
            self.json_response({'error': str(e)})
    
    def json_response(self, data):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode('utf-8'))
    
    def log_message(self, format, *args):
        print(f"[HTTP] {args[0]}")


def parse_excel(data):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    sheet_names = wb.sheetnames
    ws_name = None
    for name in sheet_names:
        if '库位' in name or '导出' in name:
            ws_name = name
            break
    if ws_name is None:
        ws_name = sheet_names[0]
    ws = wb[ws_name]
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))
    
    xs, ys, zs = set(), set(), set()
    layer_configs = {}
    shielded = []
    cell_data = {}  # (x,y,z) => {loc_id, loc_identifier, loc_name, container}
    
    # 解析第一行的仓库名和分区名
    warehouse_name = ''
    zone_name = ''
    if all_rows:
        r1 = all_rows[0]
        if r1[0] and '仓库名' in str(r1[0]):
            warehouse_name = str(r1[0])
        if len(r1) > 3 and r1[3] and '分区名' in str(r1[3]):
            zone_name = str(r1[3])
    
    # 从第三行开始处理数据行（跳过前两行表头）
    data_rows = all_rows[2:] if len(all_rows) > 2 else []
    
    for row in data_rows:
        if row[0] is None:
            continue
        
        x, y, z = None, None, None
        
        id_val = str(row[0]).strip()
        parts = id_val.split('-')
        if len(parts) >= 5:
            # 编号格式: 100-215-{X}-{Y}-{Z}
            x = int(parts[2])  # X=竖向1-12
            y = int(parts[3])  # Y=横向1-17
            z = int(parts[4])  # Z=层1-3
        
        name_val = str(row[2]).strip() if row[2] else ''
        
        # 从名称中提取（格式: {X}XY{Y}Z{Z}Z）
        m = re.match(r'(\d+)', name_val)
        if m and x is None:
            x = int(m.group(1))
        
        m = re.search(r'X(\d+)', name_val, re.IGNORECASE)
        if m and y is None:
            y = int(m.group(1))
        
        m = re.search(r'Y(\d+)', name_val, re.IGNORECASE)
        if m and z is None:
            z = int(m.group(1))
        
        if x is None or y is None:
            continue
        
        if z is None:
            z = 1
        
        xs.add(x)
        ys.add(y)
        zs.add(z)
        
        # 保存原始数据用于导出
        cell_data[(x, y, z)] = {
            'loc_id': id_val,
            'loc_identifier': str(row[1]).strip() if row[1] else id_val,
            'loc_name': name_val if name_val else f"{x}X{y}Y{z}Z",
            'container': str(row[3]).strip() if row[3] else '卡板',
        }
        
        # 记录每层的层高和承重（在第一层循环中完成，无需第二次遍历）
        if z not in layer_configs:
            height = row[5] if len(row) > 5 and row[5] is not None else ''
            weight = row[6] if len(row) > 6 and row[6] is not None else ''
            layer_configs[z] = {'height': height, 'weight': weight}
        
        # 处理所有层的屏蔽状态
        status = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        if '不可用' in status:
            shielded.append(f"{x},{y},{z}")
    
    total = len(xs) * len(ys) * len(zs)
    cfg_json = {str(k): v for k, v in layer_configs.items()}
    
    return {
        'x_range': sorted(xs),
        'y_range': sorted(ys),
        'z_range': sorted(zs),
        'total': total,
        'layer_configs': cfg_json,
        'shielded': shielded,
        'warehouse_name': warehouse_name,
        'zone_name': zone_name,
        'cell_data': {f"{k[0]},{k[1]},{k[2]}": v for k, v in cell_data.items()}
    }


def generate_excel(x_range, y_range, z_range, shielded_set, layer_configs, cell_data=None, warehouse_name='', zone_name='', cell_configs=None):
    from openpyxl.styles import Alignment, Font, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "库位导出信息"
    
    # 表头行 - 用导入的仓库名和分区名
    ws.append([warehouse_name or '仓库名：', None, None, zone_name or '分区名：', None, None, None])
    ws.append(['库位编号', '库位标识(必填)', '库位名', '适用容器', '可用状态', '层高（m）', '最大可承受重量（kg）'])
    
    # 合并单元格（A1:C1 和 D1:G1）
    ws.merge_cells('A1:C1')
    ws.merge_cells('D1:G1')
    
    # 居中对齐样式
    center = Alignment(horizontal='center', vertical='center')
    
    # 表头字体
    header_font = Font(bold=True)
    
    # 设置行高
    ws.row_dimensions[1].height = 13.5
    
    # 设置列宽
    ws.column_dimensions['A'].width = 28.6
    ws.column_dimensions['B'].width = 15.1
    ws.column_dimensions['C'].width = 16.3
    ws.column_dimensions['D'].width = 12.9
    ws.column_dimensions['E'].width = 14.0
    ws.column_dimensions['F'].width = 12.9
    ws.column_dimensions['G'].width = 21.4
    
    # 表头行样式
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.alignment = center
            if cell.row == 2:
                cell.font = header_font
    
    # 数据行
    row_num = 3
    for x in x_range:
        for y in y_range:
            for z in z_range:
                cd = cell_data.get(f"{x},{y},{z}", {}) if cell_data else {}
                loc_id = cd.get('loc_id', f"100-215-{x}-{y}-{z}")
                loc_identifier = cd.get('loc_identifier', loc_id)
                loc_name = cd.get('loc_name', f"{x}X{y}Y{z}Z")
                container = cd.get('container', '卡板')
                status = '不可用' if (x, y, z) in shielded_set else '正常'
                height = layer_configs.get(z, {}).get('height', '')
                weight = layer_configs.get(z, {}).get('weight', '')
                # 优先用 cell_configs 修改值
                if cell_configs:
                    cell_cfg = cell_configs.get(f"{x},{y},{z}", {})
                    if cell_cfg.get('height') is not None:
                        height = cell_cfg['height']
                    if cell_cfg.get('weight') is not None:
                        weight = cell_cfg['weight']
                ws.append([loc_id, loc_identifier, loc_name, container, status, height, weight])
                # 设置数据行样式
                for cell in ws[row_num]:
                    cell.alignment = center
                row_num += 1
    
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == '__main__':
    print(f"🦀 库位编辑器 v6.8")
    print(f"📍 https://0.0.0.0:{PORT}")
    class ThreadedHTTPServer(http.server.HTTPServer):
        def process_request(self, request, client_address):
            t = threading.Thread(target=self.process_request_thread, args=(request, client_address))
            t.daemon = True
            t.start()

        def process_request_thread(self, request, client_address):
            try:
                self.finish_request(request, client_address)
            except Exception:
                self.handle_error(request, client_address)
            finally:
                self.shutdown_request(request)

    server = ThreadedHTTPServer(('0.0.0.0', PORT), Handler)
    server.socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    server.max_content_length = 50 * 1024 * 1024  # 50MB upload limit

    # 加载 SSL 证书
    cert_path = '/root/certs/cert.pem'
    key_path = '/root/certs/key.pem'
    if os.path.exists(cert_path) and os.path.exists(key_path):
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
        ctx.load_cert_chain(cert_path, key_path)
        server.socket = ctx.wrap_socket(server.socket, server_side=True)
        print(f"🔒 HTTPS 模式 (自签名证书)")
    else:
        print(f"⚠️ 证书不存在，使用 HTTP 模式")
    
    server.serve_forever()
